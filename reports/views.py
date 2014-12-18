#encoding: utf-8

from django.shortcuts import redirect, render_to_response, HttpResponse
from django.core.urlresolvers import reverse
from django.template import RequestContext

from models import Report, ReportSection, ReportSectionParam
from forms import ReportForm, ReportSectionForm

from polls.models import Person, Question, Carreer
from stats import *

from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import permission_required, login_required


def admin_login(request):

	if request.user.is_authenticated():
		print request.user

		return redirect(reverse('report_home'))

	if request.method == 'GET':
		form = AuthenticationForm(request)

		return render_to_response(
			'login.html',
			{'form':form},
      		context_instance=RequestContext(request)
	    )

	elif request.method == 'POST':

		form = AuthenticationForm(request.POST)

		username = request.POST.get('username')
		password = request.POST.get('password')

		form.fields['username'].initial = username

		if not username:
			form_err = {'username_errors':'Ingresa un nombre de usuario'}
			return render_to_response(
				'login.html',
				{'form':form, 'form_err':form_err},
				context_instance=RequestContext(request)
			)
		if not password:
			form_err = {'password_errors':'Ingresa una contraseña'}
			return render_to_response(
				'login.html',
				{'form':form, 'form_err':form_err},
				context_instance=RequestContext(request)
			)

		user = authenticate(username=username, password=password)

		if user is not None:
			if user.is_active:
				login(request, user)
				return redirect(reverse('report_home'))
			else:
				form_err = {}
				form_err['error'] = 'Esta cuenta está deshabilitada'

				return render_to_response(
					'login.html',
					{'form':form, 'form_err':form_err},
					context_instance=RequestContext(request)
				)
		else:
			form_err = {}
			form_err['error'] = 'Esta cuenta no existe o la contraseña no coincide'
			return render_to_response(
				'login.html',
				{'form':form, 'form_err':form_err},
				context_instance=RequestContext(request)
			)

def admin_logout(request):
	if not request.user.is_authenticated():
		return redirect('login')

	logout(request)
	return redirect('login')


@login_required(login_url='login')
def home(request):
	reports = Report.objects.all()
	return render_to_response('reports_index.html', {'reports': reports})

@login_required(login_url='login')
def create_report(request):
	if request.method == 'GET':
		form = ReportForm()
		return render_to_response('add_edit_report.html', {'form': form}, context_instance=RequestContext(request))

	if request.method == 'POST':
		form = ReportForm(request.POST)
		if form.is_valid():
			new_report = form.save()
			return redirect(reverse('edit_report', args=(new_report.id,)))

		return render_to_response('add_edit_report.html', {'form': form}, context_instance=RequestContext(request))

@login_required(login_url='login')
def edit_report(request, report_id):

	report = Report.objects.get(id=report_id)
	
	if request.method == 'GET':
		form = ReportForm(instance=report)

		return render_to_response('add_edit_report.html', 
			{'form':form, 'editing':True, 'report':report}, 
			context_instance=RequestContext(request)
		)

	if request.method == 'POST':
		form = ReportForm(request.POST, instance=report)
		if form.is_valid():
			report = form.save()
			return redirect(reverse('report_home'))

		return render_to_response('add_edit_report.html', 
			{'form':form, 'editing':True, 'report':report}, 
			context_instance=RequestContext(request)
		)

@login_required(login_url='login')
def delete_report(request, report_id):
	report = Report.objects.get(id=report_id)
	report.delete()

	return redirect(reverse('report_home'))

@login_required(login_url='login')
def add_section(request, report_id):
	
	if request.method == 'GET':
		form = ReportSectionForm()
		return render_to_response('add_edit_section.html', {'form': form, 'report_id':report_id}, context_instance=RequestContext(request))

	if request.method == 'POST':
		form = ReportSectionForm(request.POST)
		if form.is_valid():
			new_section = form.save()

			#if new_section.attribute == 'career':
			#	return redirect(reverse('add_params', args=(new_section.id,1,)))

			return redirect(reverse('edit_report', args=(report_id,)))

		return render_to_response('add_edit_section.html', {'form': form, 'report_id':report_id}, context_instance=RequestContext(request))

"""
def add_params(request, section_id, param_id):
	
	if request.method == 'GET':
		data = {}
		if param_id == '1':
			carreers = Carreer.objects.all()
			data = {'label':'Escoja las carreras a mostrar', 'options':carreers}
			print data

		return render_to_response('add_edit_params.html', data, context_instance=RequestContext(request))

	if request.method == 'POST':
		param_string = ""

		for param in request.POST.getlist('params'):
			param_string = param_string + "," + param
			print param

		print param_string

		new_param = ReportSectionParam(
			section=ReportSection.objects.get(id=section_id),
			param_string=param_string
		)

		return render_to_response('add_edit_params.html', , context_instance=RequestContext(request))
"""

@login_required(login_url='login')
def edit_section(request, report_id, section_id):
	
	section = ReportSection.objects.get(id=section_id)

	if request.method == 'GET':
		form = ReportSectionForm(instance=section)
		return render_to_response('add_edit_section.html', 
			{'form': form, 'editing':True,'report_id':report_id, 'section_id':section_id}, 
			context_instance=RequestContext(request)
		)

	if request.method == 'POST':
		form = ReportSectionForm(request.POST, instance=section)
		if form.is_valid():
			section = form.save()
			return redirect(reverse('edit_report', args=(report_id,)))

		return render_to_response('add_edit_section.html', 
			{'form': form, 'editing':True,'report_id':report_id, 'section_id':section_id}, 
			context_instance=RequestContext(request)
		)

@login_required(login_url='login')
def delete_section(request, report_id, section_id):

	section = ReportSection.objects.get(id=section_id)
	section.delete()
	return redirect(reverse('edit_report', args=(report_id,)))

@login_required(login_url='login')
def report_observ(request, report_id):

	report = Report.objects.get(id=report_id)

	section_reports = []
	for section in report.reportsection_set.all():
		section_reports.append({
			'section': section,
			'section_report': get_section_report(section)
		})


	return render_to_response(
		'report.html', 
		{'report':report,'section_reports':section_reports},
		context_instance=RequestContext(request)
	)

def get_section_report(section):

	if section.attribute == 'genre':
		return get_stats_by_genre(section.questions.all())
	elif section.attribute == 'birth_date':
		return get_stats_by_birth(section.questions.all())
	elif section.attribute == '':
		pass

	elif section.attribute == 'career':
		carreer_ids = Person.objects.values_list(
			'career_id'
		).annotate(counter=Count('career_id')).order_by('-counter')[:5]

		ids = tuple(id[0] for id in carreer_ids)

		return get_stats_by_carreer(section.questions.all(), ids)

	elif section.attribute == 'children':
		return get_stats_by_children(section.questions.all())

	elif section.attribute == 'status':
		return get_stats_by_status(section.questions.all())

	elif section.attribute == 'stratum':
		return get_stats_by_stratum(section.questions.all())

	elif section.attribute == 'role':
		return get_stats_by_role(section.questions.all())

	elif section.attribute == 'semester':
		return get_stats_by_semester(section.questions.all())

	elif section.attribute == 'base':
		return get_stats_by_base(section.questions.all())

def export_data(request):
    import openpyxl
    from openpyxl.cell import get_column_letter

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=resultados_encuesta.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = u"Datos-Esc. de ideación suicida"

    row_num = 0

    columns = [
        (u"Encuesta #", 15),
    ]+[(u"Pregunta %s" % (question), 20) for question in Question.objects.all().values_list('index').order_by('index')]

    print columns


    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    queryset = Answer.objects.filter(application_id=1)

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.option.question.text,
            obj.option.text,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            #c.style.alignment.wrap_text = True

    wb.save(response)
    return response


